
import openpyxl
import streamlit as st


@st.cache_data
def load_8760_data(filename, month=None):
    """
    Load hourly load, solar, and wind data from the 8760 sheet.

    Args:
        filename (str): Excel file name
        month (int or None): 1-12 for one month, or None for full year

    Returns:
        tuple: (datetimes, load_vals, solar_vals, wind_vals)
    """
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb["8760"]

    datetimes = []
    load_vals = []
    solar_vals = []
    wind_vals = []

    for row in range(2, ws.max_row + 1):
        dt = ws[f"A{row}"].value
        load = ws[f"B{row}"].value
        solar = ws[f"C{row}"].value
        wind = ws[f"D{row}"].value

        if dt is None:
            continue

        if month is not None and dt.month != month:
            continue

        datetimes.append(dt)
        load_vals.append(float(load))
        solar_vals.append(0.0 if solar is None else float(solar))
        wind_vals.append(0.0 if wind is None else float(wind))

    return datetimes, load_vals, solar_vals, wind_vals
